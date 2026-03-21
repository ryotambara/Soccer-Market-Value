"""
export/excel_export.py

Generates a polished, analysis-ready Excel workbook from data/processed/results.csv.

Sheets:
  1. Players          — full player data with conditional formatting
  2. Nationality Summary — aggregated stats by nationality group + bar chart
  3. Position Summary    — aggregated stats by position group + bar chart
  4. Model Info          — static metadata + R² from model_summary.txt

Run: python export/excel_export.py
"""

import os
import re
import pandas as pd
import numpy as np
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PROC_PL_2526   = os.path.join(BASE_DIR, "data", "processed", "premier_league", "2025-26")
RESULTS_PATH    = os.path.join(_PROC_PL_2526, "results.csv")
SUMMARY_PATH    = os.path.join(_PROC_PL_2526, "model_summary.txt")
OUT_DIR         = os.path.join(BASE_DIR, "data", "exports")
OUT_PATH        = os.path.join(OUT_DIR, "pitchiq_players.xlsx")

# ── Colour constants ───────────────────────────────────────────────────────
NAVY        = "0D1420"
AMBER_TEXT  = "C47D00"
AMBER_FILL  = "FEF9EC"
RED_TEXT    = "C0392B"
RED_FILL    = "FDEDEC"
GREEN_TEXT  = "1E8449"
GREEN_FILL  = "EAFAF1"
GREY_ROW    = "F7F7F7"
WHITE       = "FFFFFF"

# ── Column widths for Sheet 1 ──────────────────────────────────────────────
COL_WIDTHS = {
    "Rank":                8,
    "Player":              24,
    "Club":                20,
    "Nationality":         18,
    "Nationality Group":   18,
    "Position":            22,
    "Position Group":      16,
    "Age":                 8,
    "Minutes Played":      14,
    "Goals / 90":          12,
    "Assists / 90":        12,
    "Contract Months":     16,
    "Team League Position":14,
    "Actual Value (€)":    18,
    "Predicted Value (€)": 18,
    "Residual":            12,
    "Valuation":           16,
    "Percentile":          12,
}


# ── Helpers ────────────────────────────────────────────────────────────────

def make_header_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=NAVY)

def make_header_font() -> Font:
    return Font(bold=True, color=WHITE, name="Calibri", size=11)

def make_thin_border() -> Border:
    s = Side(style="thin", color="E0E0E0")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def apply_header_row(ws, headers: list[str]) -> None:
    """Write bold white-on-navy header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = make_header_font()
        cell.fill = make_header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def set_col_widths(ws, width_map: dict[str, int], headers: list[str]) -> None:
    """Set column widths by header name."""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = width_map.get(header, 14)
        ws.column_dimensions[col_letter].width = width

def alternate_row_fill(ws, n_rows: int, n_cols: int, start_row: int = 2) -> None:
    """Apply alternating white / light-grey fill to data rows."""
    white_fill = make_fill(WHITE)
    grey_fill  = make_fill(GREY_ROW)
    for row_idx in range(start_row, start_row + n_rows):
        fill = grey_fill if (row_idx % 2 == 0) else white_fill
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def infer_nationality_group(row: pd.Series) -> str:
    """Reconstruct nationality group label from dummy columns."""
    groups = {
        "Brazilian":          "is_brazilian",
        "French":             "is_french",
        "English":            "is_english",
        "Spanish":            "is_spanish",
        "German":             "is_german",
        "Argentinian":        "is_argentinian",
        "Portuguese":         "is_portuguese",
        "African":            "is_african",
        "Asian":              "is_asian",
        "South American":     "is_south_american_other",
    }
    for label, col in groups.items():
        if col in row.index and row[col] == 1:
            return label
    return "Other Europe"


def infer_position_group(row: pd.Series) -> str:
    """Reconstruct position group label from dummy columns."""
    groups = {
        "Striker":       "is_striker",
        "Winger":        "is_winger",
        "Attacking Mid": "is_attacking_mid",
        "Central Mid":   "is_central_mid",
        "Fullback":      "is_fullback",
        "Goalkeeper":    "is_goalkeeper",
    }
    for label, col in groups.items():
        if col in row.index and row[col] == 1:
            return label
    return "Centre-Back"


def parse_model_stats(summary_path: str) -> dict[str, str]:
    """Extract R² and Adj. R² from the statsmodels summary text file."""
    stats = {"R²": "N/A", "Adj R²": "N/A"}
    if not os.path.exists(summary_path):
        print(f"  WARNING: model_summary.txt not found at {summary_path}. Using N/A.")
        return stats
    with open(summary_path, "r", encoding="utf-8") as f:
        text = f.read()
    # statsmodels formats: "R-squared:                       0.7654"
    m = re.search(r"R-squared:\s+([\d.]+)", text)
    if m:
        stats["R²"] = m.group(1)
    m = re.search(r"Adj\.\s+R-squared:\s+([\d.]+)", text)
    if m:
        stats["Adj R²"] = m.group(1)
    return stats


def add_bar_chart(
    ws,
    data_start_row: int,
    data_end_row: int,
    category_col: int,
    value_col: int,
    title: str,
    chart_anchor: str,
) -> None:
    """
    Add a clustered bar chart to ws.
    Categories = category_col, Values = value_col (both 1-indexed).
    data_start_row is the first DATA row (no header).
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = "Avg Residual (log)"
    chart.x_axis.title = None
    chart.style = 10
    chart.width = 18
    chart.height = 12

    n_rows = data_end_row - data_start_row + 1

    data_ref = Reference(
        ws,
        min_col=value_col,
        min_row=data_start_row - 1,  # include header for series label
        max_row=data_end_row,
    )
    cats_ref = Reference(
        ws,
        min_col=category_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Colour series bar by sign — openpyxl doesn't support per-bar colour,
    # so we use a single amber colour for the series; the conditional
    # formatting on the sheet handles the cell-level colouring.
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "F0A500"

    ws.add_chart(chart, chart_anchor)


# ── Sheet 1: Players ───────────────────────────────────────────────────────

def write_players_sheet(ws, df: pd.DataFrame) -> None:
    print("Writing Sheet 1: Players...")

    headers = list(COL_WIDTHS.keys())

    apply_header_row(ws, headers)

    # Build column values
    rows_data = []
    for rank, (_, row) in enumerate(df.iterrows(), start=1):
        nat_group  = infer_nationality_group(row)
        pos_group  = infer_position_group(row)
        percentile = int(row.get("percentile", 0))
        rows_data.append([
            rank,
            str(row.get("player_name", row.get("player", ""))),
            str(row.get("club", "")),
            str(row.get("nationality", "")),
            nat_group,
            str(row.get("position", "")),
            pos_group,
            int(row["age"]) if pd.notna(row.get("age")) else "",
            int(row["minutes_played"]) if pd.notna(row.get("minutes_played")) else "",
            round(float(row["goals_per_90"]), 2) if pd.notna(row.get("goals_per_90")) else 0.0,
            round(float(row["assists_per_90"]), 2) if pd.notna(row.get("assists_per_90")) else 0.0,
            int(row["contract_months_remaining"]) if pd.notna(row.get("contract_months_remaining")) else "",
            int(row["team_league_position"]) if pd.notna(row.get("team_league_position")) else "",
            float(row["market_value_eur"]) if pd.notna(row.get("market_value_eur")) else 0,
            float(row["predicted_market_value_eur"]) if pd.notna(row.get("predicted_market_value_eur")) else 0,
            round(float(row["residual"]), 4) if pd.notna(row.get("residual")) else 0.0,
            str(row.get("valuation_label", "")),
            f'{percentile}th',
        ])

    n_data_rows = len(rows_data)
    n_cols = len(headers)

    # Write data rows
    for r_idx, row_vals in enumerate(rows_data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Alternate row fill
    alternate_row_fill(ws, n_data_rows, n_cols, start_row=2)

    # Column-specific number formats
    header_to_col = {h: i+1 for i, h in enumerate(headers)}

    # Minutes — comma separator
    min_col = get_column_letter(header_to_col["Minutes Played"])
    for r in range(2, n_data_rows + 2):
        ws[f"{min_col}{r}"].number_format = "#,##0"

    # Currency cols
    for h in ("Actual Value (€)", "Predicted Value (€)"):
        col_l = get_column_letter(header_to_col[h])
        for r in range(2, n_data_rows + 2):
            ws[f"{col_l}{r}"].number_format = '€#,##0'
            ws[f"{col_l}{r}"].alignment = Alignment(horizontal="right", vertical="center")

    # Goals/90 and Assists/90 — 2dp
    for h in ("Goals / 90", "Assists / 90"):
        col_l = get_column_letter(header_to_col[h])
        for r in range(2, n_data_rows + 2):
            ws[f"{col_l}{r}"].number_format = "0.00"

    # Residual — always show sign
    res_col_idx = header_to_col["Residual"]
    res_col_l = get_column_letter(res_col_idx)
    for r in range(2, n_data_rows + 2):
        ws[f"{res_col_l}{r}"].number_format = '+0.000;-0.000;0.000'
        ws[f"{res_col_l}{r}"].alignment = Alignment(horizontal="center", vertical="center")

    # Rank + Age + Contract + Team Rank — center
    for h in ("Rank", "Age", "Contract Months", "Team League Position"):
        col_l = get_column_letter(header_to_col[h])
        for r in range(2, n_data_rows + 2):
            ws[f"{col_l}{r}"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Conditional formatting on Residual column ──
    res_range = f"{res_col_l}2:{res_col_l}{n_data_rows + 1}"

    # Positive → amber bold
    ws.conditional_formatting.add(
        res_range,
        CellIsRule(
            operator="greaterThan",
            formula=["0.149"],
            font=Font(color=AMBER_TEXT, bold=True),
        ),
    )
    # Negative → red bold
    ws.conditional_formatting.add(
        res_range,
        CellIsRule(
            operator="lessThan",
            formula=["-0.149"],
            font=Font(color=RED_TEXT, bold=True),
        ),
    )
    # Near zero → green bold
    ws.conditional_formatting.add(
        res_range,
        CellIsRule(
            operator="between",
            formula=["-0.149", "0.149"],
            font=Font(color=GREEN_TEXT, bold=True),
        ),
    )

    # ── Conditional formatting on Valuation column ──
    val_col_idx = header_to_col["Valuation"]
    val_col_l = get_column_letter(val_col_idx)
    val_range = f"{val_col_l}2:{val_col_l}{n_data_rows + 1}"

    ws.conditional_formatting.add(
        val_range,
        CellIsRule(
            operator="equal",
            formula=['"undervalued"'],
            font=Font(color=AMBER_TEXT, bold=True),
            fill=make_fill(AMBER_FILL),
        ),
    )

    ws.conditional_formatting.add(
        val_range,
        CellIsRule(
            operator="equal",
            formula=['"overvalued"'],
            font=Font(color=RED_TEXT, bold=True),
            fill=make_fill(RED_FILL),
        ),
    )

    ws.conditional_formatting.add(
        val_range,
        CellIsRule(
            operator="equal",
            formula=['"fairly valued"'],
            font=Font(color=GREEN_TEXT, bold=True),
            fill=make_fill(GREEN_FILL),
        ),
    )

    # ── Freeze panes: row 1 + column A (player name is col B, but freeze at B2) ──
    # Freeze top row AND left column (Rank col). To keep Player visible, freeze col C.
    ws.freeze_panes = "C2"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_data_rows + 1}"

    # ── Column widths ──
    set_col_widths(ws, COL_WIDTHS, headers)

    # ── Row height ──
    for r in range(1, n_data_rows + 2):
        ws.row_dimensions[r].height = 18

    print(f"  Written {n_data_rows} player rows.")


# ── Sheet 2: Nationality Summary ──────────────────────────────────────────

def write_nationality_sheet(ws, df: pd.DataFrame) -> None:
    print("Writing Sheet 2: Nationality Summary...")

    # Compute nationality group for each player
    df = df.copy()
    df["_nat_group"] = df.apply(infer_nationality_group, axis=1)

    headers = [
        "Nationality Group", "Player Count", "Avg Residual",
        "Avg Actual Value (€)", "Avg Predicted Value (€)",
        "Most Undervalued Player", "Most Overvalued Player",
        "% Undervalued", "% Overvalued",
    ]

    apply_header_row(ws, headers)

    groups = []
    for nat_group, grp in df.groupby("_nat_group"):
        player_col = "player_name" if "player_name" in df.columns else "player"
        most_under = grp.loc[grp["residual"].idxmax(), player_col] if not grp.empty else ""
        most_over  = grp.loc[grp["residual"].idxmin(), player_col] if not grp.empty else ""
        pct_under  = round(100 * (grp["residual"] > 0.15).sum() / len(grp), 1)
        pct_over   = round(100 * (grp["residual"] < -0.15).sum() / len(grp), 1)
        groups.append({
            "Nationality Group": nat_group,
            "Player Count": len(grp),
            "Avg Residual": round(grp["residual"].mean(), 4),
            "Avg Actual Value (€)": round(grp["market_value_eur"].mean(), 0),
            "Avg Predicted Value (€)": round(grp["predicted_market_value_eur"].mean(), 0),
            "Most Undervalued Player": most_under,
            "Most Overvalued Player": most_over,
            "% Undervalued": pct_under,
            "% Overvalued": pct_over,
        })

    groups.sort(key=lambda x: x["Avg Residual"], reverse=True)

    for r_idx, row in enumerate(groups, start=2):
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
            if h in ("Avg Actual Value (€)", "Avg Predicted Value (€)"):
                cell.number_format = "€#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif h == "Avg Residual":
                cell.number_format = "+0.00;-0.00;0.00"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    color=AMBER_TEXT if row[h] > 0 else RED_TEXT if row[h] < 0 else GREEN_TEXT,
                    bold=True,
                )
            elif h in ("% Undervalued", "% Overvalued"):
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif h == "Player Count":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    n_data_rows = len(groups)
    alternate_row_fill(ws, n_data_rows, len(headers), start_row=2)

    # Re-apply coloured font on Avg Residual after alternate fill (fill overwrites)
    res_col_l = get_column_letter(3)  # "Avg Residual" is column 3
    for r_idx, row in enumerate(groups, start=2):
        cell = ws[f"{res_col_l}{r_idx}"]
        val = row["Avg Residual"]
        cell.font = Font(
            color=AMBER_TEXT if val > 0 else RED_TEXT if val < 0 else GREEN_TEXT,
            bold=True,
        )

    # Column widths
    nat_widths = {h: 22 for h in headers}
    nat_widths["Player Count"] = 14
    nat_widths["Avg Residual"] = 16
    nat_widths["% Undervalued"] = 15
    nat_widths["% Overvalued"] = 15
    set_col_widths(ws, nat_widths, headers)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_data_rows + 1}"

    # ── Bar chart ──
    add_bar_chart(
        ws,
        data_start_row=2,
        data_end_row=n_data_rows + 1,
        category_col=1,  # Nationality Group
        value_col=3,      # Avg Residual
        title="Average Valuation Residual by Nationality",
        chart_anchor="K2",
    )

    print(f"  Written {n_data_rows} nationality groups.")


# ── Sheet 3: Position Summary ─────────────────────────────────────────────

def write_position_sheet(ws, df: pd.DataFrame) -> None:
    print("Writing Sheet 3: Position Summary...")

    df = df.copy()
    df["_pos_group"] = df.apply(infer_position_group, axis=1)

    headers = [
        "Position Group", "Player Count", "Avg Residual",
        "Avg Goals / 90", "Avg Assists / 90",
        "Avg Actual Value (€)", "Avg Predicted Value (€)",
        "Most Undervalued Player", "Most Overvalued Player",
    ]

    apply_header_row(ws, headers)

    groups = []
    for pos_group, grp in df.groupby("_pos_group"):
        player_col = "player_name" if "player_name" in df.columns else "player"
        most_under = grp.loc[grp["residual"].idxmax(), player_col] if not grp.empty else ""
        most_over  = grp.loc[grp["residual"].idxmin(), player_col] if not grp.empty else ""
        groups.append({
            "Position Group": pos_group,
            "Player Count": len(grp),
            "Avg Residual": round(grp["residual"].mean(), 4),
            "Avg Goals / 90": round(grp["goals_per_90"].mean(), 3) if "goals_per_90" in grp else 0,
            "Avg Assists / 90": round(grp["assists_per_90"].mean(), 3) if "assists_per_90" in grp else 0,
            "Avg Actual Value (€)": round(grp["market_value_eur"].mean(), 0),
            "Avg Predicted Value (€)": round(grp["predicted_market_value_eur"].mean(), 0),
            "Most Undervalued Player": most_under,
            "Most Overvalued Player": most_over,
        })

    groups.sort(key=lambda x: x["Avg Residual"], reverse=True)

    for r_idx, row in enumerate(groups, start=2):
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[h])
            if h in ("Avg Actual Value (€)", "Avg Predicted Value (€)"):
                cell.number_format = "€#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif h == "Avg Residual":
                cell.number_format = "+0.00;-0.00;0.00"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    color=AMBER_TEXT if row[h] > 0 else RED_TEXT if row[h] < 0 else GREEN_TEXT,
                    bold=True,
                )
            elif h in ("Avg Goals / 90", "Avg Assists / 90"):
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif h == "Player Count":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    n_data_rows = len(groups)
    alternate_row_fill(ws, n_data_rows, len(headers), start_row=2)

    # Re-apply coloured font on Avg Residual
    res_col_l = get_column_letter(3)
    for r_idx, row in enumerate(groups, start=2):
        cell = ws[f"{res_col_l}{r_idx}"]
        val = row["Avg Residual"]
        cell.font = Font(
            color=AMBER_TEXT if val > 0 else RED_TEXT if val < 0 else GREEN_TEXT,
            bold=True,
        )

    pos_widths = {h: 22 for h in headers}
    pos_widths["Player Count"] = 14
    pos_widths["Avg Residual"] = 16
    pos_widths["Avg Goals / 90"] = 16
    pos_widths["Avg Assists / 90"] = 16
    set_col_widths(ws, pos_widths, headers)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_data_rows + 1}"

    add_bar_chart(
        ws,
        data_start_row=2,
        data_end_row=n_data_rows + 1,
        category_col=1,
        value_col=3,
        title="Average Valuation Residual by Position Group",
        chart_anchor="K2",
    )

    print(f"  Written {n_data_rows} position groups.")


# ── Sheet 4: Model Info ───────────────────────────────────────────────────

def write_model_info_sheet(ws, df: pd.DataFrame, model_stats: dict[str, str]) -> None:
    print("Writing Sheet 4: Model Info...")

    today_str = date.today().strftime("%d %b %Y")
    n_players = len(df)

    rows = [
        ("Dataset",           "Premier League 2024-25"),
        ("Players Included",  str(n_players)),
        ("Min Minutes",       "500"),
        ("Model Type",        "OLS Regression (statsmodels)"),
        ("Dependent Var",     "log(Market Value)"),
        ("R²",                model_stats.get("R²", "N/A")),
        ("Adj R²",            model_stats.get("Adj R²", "N/A")),
        ("Generated On",      today_str),
        ("Data Sources",      "Transfermarkt, FBref"),
        ("Baseline Position", "Centre-Back"),
        ("Baseline Nation",   "Other Europe"),
    ]

    # Header row
    ws.cell(row=1, column=1, value="Parameter").font = make_header_font()
    ws.cell(row=1, column=1).fill = make_header_fill()
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=1, column=2, value="Value").font = make_header_font()
    ws.cell(row=1, column=2).fill = make_header_fill()
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, (param, value) in enumerate(rows, start=2):
        p_cell = ws.cell(row=r_idx, column=1, value=param)
        p_cell.font = Font(bold=True, name="Calibri", size=11)
        p_cell.alignment = Alignment(horizontal="left", vertical="center")
        p_cell.fill = make_fill(WHITE if r_idx % 2 != 0 else GREY_ROW)

        v_cell = ws.cell(row=r_idx, column=2, value=value)
        v_cell.alignment = Alignment(horizontal="left", vertical="center")
        v_cell.fill = make_fill(WHITE if r_idx % 2 != 0 else GREY_ROW)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.row_dimensions[1].height = 20

    print(f"  Written {len(rows)} info rows.")


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    print("=" * 60)
    print("PitchIQ — Excel Export")
    print("=" * 60)

    # Load results
    if not os.path.exists(RESULTS_PATH):
        print(f"ERROR: {RESULTS_PATH} not found. Run model/regression.py first.")
        return

    print(f"\nLoading {RESULTS_PATH}...")
    df = pd.read_csv(RESULTS_PATH, encoding="utf-8")

    # Sort by residual descending (most undervalued first)
    df = df.sort_values("residual", ascending=False).reset_index(drop=True)

    # Ensure predicted_market_value_eur exists
    if "predicted_market_value_eur" not in df.columns:
        import numpy as np
        df["predicted_market_value_eur"] = np.exp(df["predicted_log_value"])

    # Compute percentile if missing
    if "percentile" not in df.columns:
        df["percentile"] = (df["residual"].rank(pct=True) * 100).round(0).astype(int)

    # Compute valuation_label if missing
    if "valuation_label" not in df.columns:
        def label(r):
            if r > 0.15:   return "overvalued"
            if r < -0.15:  return "undervalued"
            return "fairly valued"
        df["valuation_label"] = df["residual"].apply(label)

    print(f"  Loaded {len(df)} players.")

    # Parse model stats
    model_stats = parse_model_stats(SUMMARY_PATH)
    print(f"  Model R²: {model_stats['R²']}  Adj R²: {model_stats['Adj R²']}")

    # Create workbook
    wb = Workbook()

    # Rename default sheet
    ws1 = wb.active
    ws1.title = "Players"

    ws2 = wb.create_sheet("Nationality Summary")
    ws3 = wb.create_sheet("Position Summary")
    ws4 = wb.create_sheet("Model Info")

    print()
    write_players_sheet(ws1, df)
    print()
    write_nationality_sheet(ws2, df)
    print()
    write_position_sheet(ws3, df)
    print()
    write_model_info_sheet(ws4, df, model_stats)

    print(f"\nSaving workbook...")
    wb.save(OUT_PATH)

    print(f"\n{'=' * 60}")
    print(f"Saved to {OUT_PATH}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
